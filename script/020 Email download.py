# -*- coding: utf-8 -*-
"""
020 Email download.py  (北京时间版)
- 统一所有时间为 Asia/Shanghai（UTC+8）
- mail_meta.json 内写入带偏移的 ISO8601
"""
import os
import sys
import re
import platform
import json
import email
import imaplib
from email.header import decode_header
from email.utils import parsedate_tz, mktime_tz
from datetime import datetime
from zoneinfo import ZoneInfo  # Python 3.9+

import pandas as pd
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment
from dotenv import load_dotenv

# ================================
# 🕒 时区工具（统一北京时间）
# ================================
TZ_SH = ZoneInfo("Asia/Shanghai")


def mask_email(value: str | None) -> str:
    if not value:
        return "(empty)"
    if "@" not in value:
        return value[:2] + "***"
    name, domain = value.split("@", 1)
    if len(name) <= 2:
        name_masked = name[0] + "*"
    else:
        name_masked = name[:2] + "***"
    return f"{name_masked}@{domain}"


def mask_secret(value: str | None) -> str:
    if not value:
        return "(empty)"
    return f"***len={len(value)}"

def now_shanghai() -> datetime:
    return datetime.now(TZ_SH)

def ts_to_shanghai(ts: float) -> datetime:
    return datetime.fromtimestamp(ts, tz=TZ_SH)

# ================================
# 📂 路径配置（支持主程序传参）
# ================================
if platform.system() == "Windows":
    default_save_path = os.path.join(os.getcwd(), "data")
else:
    default_save_path = os.path.expanduser("~/data")

excel_save_path = sys.argv[1] if len(sys.argv) >= 2 else default_save_path
os.makedirs(excel_save_path, exist_ok=True)
print(f"📂 保存路径: {os.path.abspath(excel_save_path)}")

# ================================
# 🔧 关键词/邮箱配置（集中管理）
# ================================
KEYWORDS = {
    "waiting": "等待您查看",
    "heyu_da": "合肥市和裕达",
}
MAILBOX = os.getenv("IMAP_MAILBOX", "INBOX")
RECENT_LIMIT = int(os.getenv("RECENT_LIMIT", "15"))
META_FILENAME = "mail_meta.json"

# ================================
# 📧 邮箱凭据（.env）
# ================================
load_dotenv()
email_user = os.getenv("EMAIL_ADDRESS_QQ")
email_password = os.getenv("EMAIL_PASSWORD_QQ") or os.getenv("EMAIL_PASSWOR_QQ")
email_server = os.getenv("IMAP_SERVER", "imap.qq.com")

if not email_user or not email_password:
    raise ValueError("❌ 环境变量未正确配置（EMAIL_ADDRESS_QQ / EMAIL_PASSWORD_QQ）！")

print("📬 环境变量检查：")
print("   EMAIL_ADDRESS_QQ =", mask_email(email_user))
print("   EMAIL_PASSWORD_QQ/EMAIL_PASSWOR_QQ =", mask_secret(email_password))
print("   IMAP_SERVER =", email_server)
print("📬 正在使用邮箱:", mask_email(email_user))

# ================================
# 🔑 标题解码与清理
# ================================
def decode_str(s: str) -> str:
    if not s:
        return ""
    value, charset = decode_header(s)[0]
    if charset:
        value = value.decode(charset)
    elif isinstance(value, bytes):
        value = value.decode("utf-8", errors="ignore")
    return value

def clean_subject(subject: str) -> str:
    cleaned_subject = re.sub(r'\[([^\[\]]+)\]', r'\1', subject or "")
    cleaned_subject = re.sub(r'【([^【】]+)】', r'\1', cleaned_subject)
    return cleaned_subject.strip()

# ================================
# 📨 抓取邮件并输出 HTML/元数据/附件
# ================================
def fetch_html_from_emails(server: str, user: str, password: str, save_dir: str) -> str | None:
    mail = None
    html_content = None

    meta = {
        "selected_heyu_da_subject": None,
        "selected_heyu_da_received_at": None,  # ISO8601（带+08:00）
        "selected_waiting_subject": None,
        "selected_waiting_received_at": None,  # ISO8601（带+08:00）
    }

    try:
        print("🔗 正在连接邮箱...")
        mail = imaplib.IMAP4_SSL(server)
        mail.login(user, password)

        status, _ = mail.select(MAILBOX)
        if status != "OK":
            print(f"⚠️ 无法选择邮箱目录 {MAILBOX}，尝试使用 INBOX")
            mail.select("INBOX")

        print(f"🔎 正在检索最近 {RECENT_LIMIT} 封邮件...")
        status, messages = mail.search(None, "ALL")
        if status != "OK":
            print("未找到邮件")
            return None

        mail_ids = messages[0].split()
        if not mail_ids:
            print("邮箱为空。")
            return None

        recent_mail_ids = mail_ids[-RECENT_LIMIT:]
        print(f"📨 共 {len(mail_ids)} 封，处理最近 {len(recent_mail_ids)} 封。")

        inventory_query_emails = []

        for i, mail_id in enumerate(recent_mail_ids, start=1):
            status, msg_data = mail.fetch(mail_id, "(RFC822)")
            if status != "OK" or not msg_data or not msg_data[0]:
                print(f"⚠️ 第 {i} 封抓取失败")
                continue

            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)

            subject = decode_str(msg.get("Subject"))
            from_ = decode_str(msg.get("From"))
            date_raw = decode_str(msg.get("Date"))

            mail_date = parsedate_tz(date_raw)
            if mail_date:
                # mktime_tz 返回 UTC 秒数；直接转换为“北京时间” aware datetime
                mail_datetime = ts_to_shanghai(mktime_tz(mail_date))
            else:
                mail_datetime = datetime(1970, 1, 1, tzinfo=TZ_SH)

            cleaned_subject = clean_subject(subject)
            print(f"  · 第 {i} 封 | 原: {subject} | 清理: {cleaned_subject} | 发件人: {from_} | 收到(北京): {mail_datetime.strftime('%Y-%m-%d %H:%M:%S %z')}")

            if (KEYWORDS["waiting"] in cleaned_subject) or (KEYWORDS["heyu_da"] in cleaned_subject):
                inventory_query_emails.append({
                    "mail_id": mail_id,
                    "subject": subject,
                    "cleaned_subject": cleaned_subject,
                    "date": mail_datetime,  # Aware(Asia/Shanghai)
                    "msg": msg
                })

        if inventory_query_emails:
            print("\n✅ 命中关键词的邮件：")
            for item in inventory_query_emails:
                print(f"  - {item['cleaned_subject']} | {item['date'].strftime('%Y-%m-%d %H:%M:%S %z')}")
        else:
            print("\nℹ️ 未命中任何关键词邮件。")

        # 选出“合肥市和裕达”最新一封
        selected_heyu = _pick_latest(inventory_query_emails, KEYWORDS["heyu_da"])
        if selected_heyu:
            html_content = extract_html_from_msg(selected_heyu["msg"]) or html_content
            print(f"\n📌 选中(合肥市和裕达): {selected_heyu['cleaned_subject']} | {selected_heyu['date'].strftime('%Y-%m-%d %H:%M:%S %z')}")
            meta["selected_heyu_da_subject"] = selected_heyu["cleaned_subject"]
            meta["selected_heyu_da_received_at"] = selected_heyu["date"].isoformat()
            download_attachments(selected_heyu["msg"], save_dir)

        # 选出“等待您查看”最新一封
        selected_waiting = _pick_latest(inventory_query_emails, KEYWORDS["waiting"])
        if selected_waiting:
            html_content = extract_html_from_msg(selected_waiting["msg"]) or html_content
            print(f"\n📌 选中(等待您查看): {selected_waiting['cleaned_subject']} | {selected_waiting['date'].strftime('%Y-%m-%d %H:%M:%S %z')}")
            meta["selected_waiting_subject"] = selected_waiting["cleaned_subject"]
            meta["selected_waiting_received_at"] = selected_waiting["date"].isoformat()

        _write_meta(meta, os.path.join(save_dir, META_FILENAME))

        if html_content:
            print("✅ 已获取选定邮件的 HTML 正文。")
        else:
            print("ℹ️ 未找到符合条件的 HTML 正文。")

        return html_content

    except imaplib.IMAP4.error as e:
        print(f"IMAP 错误: {e}")
        return None
    except Exception as e:
        print(f"获取邮件失败: {e}")
        return None
    finally:
        try:
            if mail is not None:
                mail.logout()
        except Exception:
            pass

def _pick_latest(candidates: list[dict], keyword: str) -> dict | None:
    selected = None
    for item in candidates:
        if keyword in item["cleaned_subject"]:
            if (selected is None) or (item["date"] > selected["date"]):
                selected = item
    return selected

# ================================
# 🧩 从邮件中提取 HTML 正文
# ================================
def extract_html_from_msg(msg) -> str | None:
    html_content = None
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            content_disposition = str(part.get("Content-Disposition") or "")
            if content_type == "text/html" and "attachment" not in content_disposition:
                charset = part.get_content_charset() or part.get_charset() or "utf-8"
                try:
                    html_content = part.get_payload(decode=True).decode(charset, errors="ignore")
                except Exception:
                    html_content = part.get_payload(decode=True).decode("utf-8", errors="ignore")
                break
    else:
        if msg.get_content_type() == "text/html":
            charset = msg.get_content_charset() or msg.get_charset() or "utf-8"
            try:
                html_content = msg.get_payload(decode=True).decode(charset, errors="ignore")
            except Exception:
                html_content = msg.get_payload(decode=True).decode("utf-8", errors="ignore")
    return html_content

# ================================
# 📎 下载附件（文件名追加“北京时间”时间戳）
# ================================
def download_attachments(msg, download_folder: str) -> None:
    """下载邮件附件：文件名按 原名_YYYYMMDD_HHMMSS（北京）+扩展名。"""
    if not msg.is_multipart():
        return

    import mimetypes
    import unicodedata

    def _decode_filename(raw: str) -> str:
        parts = decode_header(raw)
        s = ""
        for p, enc in parts:
            if isinstance(p, bytes):
                s += p.decode(enc or "utf-8", errors="ignore")
            else:
                s += p
        s = unicodedata.normalize("NFC", s).replace("．", ".").strip().strip(".")
        return s

    def _sanitize(name: str) -> str:
        invalid = '<>:"/\\|?*'
        name = "".join((c if c not in invalid else "_") for c in name).strip().strip(".")
        return name or "attachment"

    def _guess_ext(content_type: str) -> str:
        overrides = {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
            "application/vnd.ms-excel": ".xls",
            "text/csv": ".csv",
            "application/zip": ".zip",
            "application/pdf": ".pdf",
        }
        return overrides.get(content_type) or (mimetypes.guess_extension(content_type) or "")

    def _ensure_unique(path: str) -> str:
        if not os.path.exists(path):
            return path
        base, ext = os.path.splitext(path)
        i = 2
        while True:
            candidate = f"{base}_{i}{ext}"
            if not os.path.exists(candidate):
                return candidate
            i += 1

    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue

        content_disposition = str(part.get("Content-Disposition") or "")
        raw_name = part.get_filename()

        if "attachment" not in content_disposition and not raw_name:
            continue

        if raw_name:
            filename = _decode_filename(raw_name)
        else:
            filename = f"attachment{_guess_ext(part.get_content_type())}"

        base_name, ext = os.path.splitext(filename)
        if not ext:
            ext = _guess_ext(part.get_content_type())

        ts = now_shanghai().strftime("%Y%m%d_%H%M%S")  # 北京时间
        safe_base = _sanitize(base_name)
        safe_name = f"{safe_base}_{ts}{ext}"
        file_path = os.path.join(download_folder, safe_name)
        file_path = _ensure_unique(file_path)

        file_data = part.get_payload(decode=True)
        if not file_data:
            continue
        with open(file_path, "wb") as f:
            f.write(file_data)

        print(f"📥 附件已下载(北京时): {file_path}")

# ================================
# 🧠 解析 HTML 表格并导出 Excel
# ================================
def parse_html_table(html_content: str) -> list[list[str]]:
    print("正在解析 HTML 内容中的表格...")

    try:
        snap_path = os.path.join(excel_save_path, "last_mail_html.html")
        with open(snap_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        print(f"🔎 HTML 快照: {snap_path}")
    except Exception:
        pass

    soup = BeautifulSoup(html_content, "html.parser")
    table = soup.find("table")
    if not table:
        print("未找到 HTML 表格！")
        return []

    data = []
    rows = table.find_all("tr")
    header = None

    for idx, row in enumerate(rows):
        cols = [ele.get_text(strip=True) for ele in row.find_all(["td", "th"])]
        if not cols:
            print(f"第 {idx + 1} 行是空行，跳过")
            continue
        if header is None:
            if idx == 0 and len(cols) > 10:
                print("第一行列数过多，认为其为正文内容，跳过")
                continue
            header = cols
            data.append(header)
            continue
        if len(cols) != len(header):
            print(f"第 {idx + 1} 行列数与表头不匹配，跳过")
            continue
        if cols == header:
            print(f"第 {idx + 1} 行是重复表头，跳过")
            continue
        data.append(cols)

    print(f"成功提取 {len(data)} 行表格数据。")

    for i in range(len(data)):
        for j in range(len(data[i])):
            if isinstance(data[i][j], str) and data[i][j].isdigit():
                data[i][j] = str(data[i][j])

    return data

def save_to_excel(data: list[list[str]], save_dir: str, file_prefix="存量查询") -> None:
    if not data:
        print("ℹ️ 没有可导出的数据。")
        return

    seen = set()
    unique_data = []
    for row in data:
        tup = tuple(row)
        if tup not in seen:
            seen.add(tup)
            unique_data.append(row)

    df = pd.DataFrame(unique_data)

    # 文件名用北京时间
    timestamp = now_shanghai().strftime("%Y%m%d_%H%M%S")
    file_name = f"{file_prefix}_{timestamp}.xlsx"
    full_path = os.path.join(save_dir, file_name)

    print(f"💾 正在保存 Excel（北京时）: {full_path}")
    df.to_excel(full_path, index=False, header=False)

    wb = openpyxl.load_workbook(full_path)
    ws = wb.active
    ws.title = "第一页"

    decimal_columns = [4, 5]  # 第5/6列（0-based）
    max_row = ws.max_row
    for col in decimal_columns:
        numeric_count = 0
        for r in range(2, max_row + 1):
            val = ws.cell(row=r, column=col + 1).value
            try:
                float(str(val).replace(",", ""))
                numeric_count += 1
            except Exception:
                pass
        if numeric_count >= (max_row - 1) / 2:
            for r in range(2, max_row + 1):
                cell = ws.cell(row=r, column=col + 1)
                try:
                    v = float(str(cell.value).replace(",", ""))
                    cell.value = v
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                except Exception:
                    pass

    wb.save(full_path)
    print("✅ Excel 保存完成。")

def _write_meta(meta: dict, path: str) -> None:
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)
        print(f"📝 元数据已写入: {path}")
    except Exception as e:
        print(f"⚠️ 元数据写入失败: {e}")

# ================================
# 🚀 主程序
# ================================
if __name__ == '__main__':
    print(f"程序启动（北京时）: {now_shanghai().strftime('%Y-%m-%d %H:%M:%S %z')}")
    html_content = fetch_html_from_emails(email_server, email_user, email_password, excel_save_path)

    if html_content:
        preview = html_content[:400].replace("\n", " ")
        print(f"HTML 预览: {preview} ...")

        table_data = parse_html_table(html_content)
        if table_data:
            save_to_excel(table_data, excel_save_path, file_prefix="存量查询")
        else:
            print("表格为空，未导出 Excel。")
    else:
        print("未获取到 HTML，程序结束。")
        raise SystemExit(1)
