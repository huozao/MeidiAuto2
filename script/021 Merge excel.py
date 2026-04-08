# -*- coding: utf-8 -*-

# ================================================
# STEP CARD
# 功能: 以最新“合肥市”文件为基底，合并为总库存主文件。
# 输入: 020 产出的 Excel
# 输出: 总库存*.xlsx
# 上游: 020 Email download.py
# 下游: 030/032/033/041/042/050
# ================================================

"""
021 Merge excel.py
- 以【最新的】文件名包含“合肥市”的 Excel 作为基底
- 排除其它所有“合肥市”旧文件，避免覆盖
- 输出文件名后缀使用北京时间（UTC+8）
"""
import os
import sys
import shutil
import platform
from datetime import datetime
from openpyxl import load_workbook
from zoneinfo import ZoneInfo   # Python 3.9+ 内置时区库

# ================================
# ⚙️ 配置区
# ================================
KEYWORD_BASE = "合肥市"
EXT = ".xlsx"
SORT_OTHERS_BY_MTIME_ASC = True
PRINT_PREFIX = "✅"

# ================================
# 📂 路径获取
# ================================
if platform.system() == "Windows":
    default_folder_path = os.path.join(os.getcwd(), "data")
else:
    default_folder_path = os.path.join(os.getcwd(), "data")

if len(sys.argv) >= 2:
    folder_path = os.path.join(sys.argv[1])
    print(f"{PRINT_PREFIX} 使用传入路径: {folder_path}")
else:
    folder_path = default_folder_path
    print(f"⚠️ 未传入路径，使用默认路径: {folder_path}")

if not os.path.exists(folder_path):
    print(f"❌ 文件夹不存在: {folder_path}")
    sys.exit(1)

# 有效文件过滤
def is_valid_xlsx(name: str) -> bool:
    return (
        name.endswith(EXT)
        and not os.path.basename(name).startswith("~$")
        and os.path.isfile(os.path.join(folder_path, name))
    )

all_excels = [f for f in os.listdir(folder_path) if is_valid_xlsx(f)]
if not all_excels:
    print("⚠️ 目录下没有可用的 .xlsx 文件")
    sys.exit(0)

# ================================
# 🔎 找最新的“合肥市”文件
# ================================
hefei_candidates = [f for f in all_excels if KEYWORD_BASE in f]
print(f"{PRINT_PREFIX} 找到 {len(hefei_candidates)} 个文件名包含 '{KEYWORD_BASE}':")
for i, f in enumerate(hefei_candidates, 1):
    print(f"{i}. {f}")

if not hefei_candidates:
    print(f"❌ 未找到包含“{KEYWORD_BASE}”的基底文件")
    sys.exit(1)

hefei_candidates = sorted(
    hefei_candidates,
    key=lambda f: os.path.getmtime(os.path.join(folder_path, f)),
    reverse=True
)
base_file = hefei_candidates[0]
base_path = os.path.join(folder_path, base_file)
print(f"\n{PRINT_PREFIX} 基底文件（最新）: {base_file}")

# ================================
# 🧺 其它待合并文件
# ================================
other_excel_files = [f for f in all_excels if f not in hefei_candidates]
if SORT_OTHERS_BY_MTIME_ASC:
    other_excel_files = sorted(other_excel_files, key=lambda f: os.path.getmtime(os.path.join(folder_path, f)))
else:
    other_excel_files = sorted(other_excel_files, key=lambda f: os.path.getmtime(os.path.join(folder_path, f)), reverse=True)

print(f"{PRINT_PREFIX} 待合并文件数: {len(other_excel_files)}")
for i, f in enumerate(other_excel_files, 1):
    print(f"{i}. {f}")

# ================================
# 🆕 创建合并文件（文件名用北京时间）
# ================================
beijing_now = datetime.now(ZoneInfo("Asia/Shanghai"))
timestamp = beijing_now.strftime("%Y%m%d_%H%M%S")
merged_filename = f"总库存{timestamp}.xlsx"
merged_filepath = os.path.join(folder_path, merged_filename)

shutil.copy(base_path, merged_filepath)
print(f"\n{PRINT_PREFIX} 创建合并文件: {merged_filename}")

# ================================
# 📑 合并逻辑
# ================================
def copy_sheet_values(src_wb, dst_wb, sheet_name: str):
    src = src_wb[sheet_name]
    if sheet_name in dst_wb.sheetnames:
        del dst_wb[sheet_name]
    dst = dst_wb.create_sheet(sheet_name)
    for row in src.iter_rows(values_only=True):
        dst.append(list(row))

merged_wb = load_workbook(merged_filepath)

for file in other_excel_files:
    file_path = os.path.join(folder_path, file)
    try:
        wb = load_workbook(file_path, data_only=True, read_only=False)
    except Exception as e:
        print(f"⚠️ 跳过无法打开的文件: {file}，原因：{e}")
        continue

    for sheet_name in wb.sheetnames:
        copy_sheet_values(wb, merged_wb, sheet_name)
        print(f"{PRINT_PREFIX} 复制工作表: {sheet_name} ← {file}")
    wb.close()

if "Sheet" in merged_wb.sheetnames and len(merged_wb["Sheet"]["A"]) == 0:
    try:
        del merged_wb["Sheet"]
    except Exception:
        pass

merged_wb.save(merged_filepath)
merged_wb.close()
print(f"{PRINT_PREFIX} 合并完成，输出: {merged_filepath}")
