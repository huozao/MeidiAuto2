# -*- coding: utf-8 -*-
"""
021 Merge excel.py
说明：
- 在库存表中插入所需列、生成“第一页副本/家里库存”、按 4 位/5 位编码把“家里库存”的数量回填到库存表 M 列，
  并逐条打印匹配详情。
- ⚠️ 将“等待您查看”的收到时间写入 M3 的动作放在**全部匹配与格式化完成之后**再执行。
"""

import os
import sys
import re
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import Selection
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from pipeline.io_utils import resolve_data_dir, find_latest_excel

# =========================================
# 🔧 CONFIG｜集中配置（只改这里）
# -----------------------------------------
CONFIG = {
    # 0) 邮件元数据（写入 M3 的时间来源）
    "meta_filename": "mail_meta.json",                  # data 目录下的 json 文件
    "meta_waiting_key": "selected_waiting_received_at", # mail_meta.json 中的键

    # 1) 路径与目标匹配
    "default_folder": os.path.join(os.getcwd(), "data"),   # 未传参时的默认目录
    "inventory_pattern": "*总库存*.xlsx",                   # 库存文件名匹配模式
    "target_sheet": "库存表",                               # 目标工作表（被写入/格式化）

    # 2) 有效数据范围与结构
    "first_data_row": 4,     # 从第几行开始视为表体（用于找B列第一个空行）
    "align_start_row": 5,    # 从第几行开始应用会计格式/对齐
    "insert_after_J_cols": 10,  # 在 J 列后插入多少列
    "insert_after_B_cols": 1,   # 在 B 列后插入多少列（用于编号列）

    # 3) 版面与视图
    "freeze_panes": "A5",    # 冻结窗格位置
    "row1_height": 18,       # 第1行行高
    "hide_row2": True,       # 是否隐藏第2行
    "zoom_scale": 95,        # 打开缩放
    "focus_cell": "A5",      # 打开时聚焦单元格（并激活 target_sheet）

    # 4) 列宽与隐藏
    "column_widths": {       # 宽度单位为Excel列宽（会在设置时 +0.6）
        "B": 4.5, "C": 0.1, "D": 35.88, "E": 3, "F": 3.6,
        "G": 8.6, "H": 8, "I": 8, "J": 8.8,
        "K": 5.88, "L": 8.1, "M": 9.8, "N": 5.88,
        "O": 9.5, "P": 9.8, "Q": 10.08, "R": 9.5, "S": 9.5, "T": 18
    },
    "hidden_columns": ["C"],     # 打开时默认隐藏的列（不影响读写）
    "left_align_cols": ["T"],    # 需要左对齐的列（列字母），例：T列左对齐

    # 5) 表头与合并
    "merge_ranges": [("H3", "J3"), ("U3", "W3")],  # 合并区域
    "set_value_cells": {"U3": "不合格"},           # 合并后需要写入的单元格内容
    "headers_K_row4": [                            # 从 K4 起向右写入的表头
        "外应存", "最小发货", "家里库存", "家应存", "排产",
        "月计划", "月计划缺口", "外仓出库总量", "外仓入库总量", "备注"
    ],
    "header_fill_color": "C187F7",                 # 表头填充色（16进制RGB）

    # 6) “第一页副本”与“家里库存”派生
    "src_sheet_for_copy": "第一页",  # 若存在则筛选生成“第一页副本”
    "warehouse_col_name": "仓库",
    "warehouse_keep_value": "成品库",
    "copy_sheet_name": "第一页副本",
    "home_sheet_name": "家里库存",
    "home_cols": ["编号", "存货名称", "数量"],  # 家里库存三列表头
    "home_name_col": "存货名称",
    "home_qty_col": "主数量",

    # 7) 回填规则（家里库存 → 库存表）
    "backfill_target_col_index": 13,  # 回填到库存表的列索引（M列=13）
    "regex_4digit_dash": r"\d{4}-",   # 'dddd-' 用后4位匹配
    "regex_5digit": r"\d{5}",         # 5位标准编号匹配

    # 8) 会计格式与对齐（G~Q）
    "acc_fmt_cols": (7, 17),  # 列范围（G=7 ~ Q=17）
    "acc_fmt_row_from": 5,    # 从第5行开始
    "acc_number_format": '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)',

    # 9) 其他
    "center_col_letter": "C",  # 需要居中的编号列
    "center_title_cell": "B1", # 左对齐的标题单元格
    "number_header_cell": "C4" # 编号列表头
}
# =========================================


def _read_waiting_time(folder, meta_name, key):
    """读取 mail_meta.json 中的 selected_waiting_received_at（字符串）"""
    meta_path = os.path.join(folder, meta_name)
    try:
        with open(meta_path, "r", encoding="utf-8") as f:
            meta = json.load(f)
        val = meta.get(key)
        return str(val) if val else None
    except FileNotFoundError:
        print(f"ℹ️ 未找到元数据文件: {meta_path}")
    except Exception as e:
        print(f"⚠️ 读取元数据失败: {e}")
    return None


def main(cfg: dict):
    # ---------- 路径 ----------
    folder_path = str(resolve_data_dir(sys.argv[1] if len(sys.argv) >= 2 else None))
    if not os.path.exists(folder_path):
        print(f"❌ 路径不存在: {folder_path}")
        sys.exit(1)

    latest_file = find_latest_excel(Path(folder_path), cfg["inventory_pattern"])
    if not latest_file:
        print("❌ 未找到包含“总库存”的文件")
        sys.exit(1)
    print(f"📄 处理文件：{latest_file}")

    # ---------- 打开工作簿 ----------
    wb = load_workbook(latest_file)
    if cfg["target_sheet"] not in wb.sheetnames:
        print(f"❌ 缺少工作表：{cfg['target_sheet']}")
        sys.exit(1)
    sh = wb[cfg["target_sheet"]]

    # ---------- 找B列第一个空行 ----------
    max_row = sh.max_row
    last_empty_row = max_row + 1
    for r in range(cfg["first_data_row"], max_row + 1):
        if sh[f"B{r}"].value is None:
            last_empty_row = r
            break
    print(f"⚡ B列第一个空行: {last_empty_row}")

    # ---------- 解除合并 ----------
    for rng in list(sh.merged_cells.ranges):
        sh.unmerge_cells(str(rng))

    # ---------- 插入列 ----------
    sh.insert_cols(10, cfg["insert_after_J_cols"])  # J 后插入
    sh.insert_cols(3,  cfg["insert_after_B_cols"])  # B 后插入（编号列）

    # ---------- 对齐 ----------
    for c in sh[cfg["center_col_letter"]]:
        c.alignment = Alignment(horizontal="center", vertical="center")
    sh[cfg["center_title_cell"]].alignment = Alignment(horizontal="left", vertical="center")

    # ---------- 合并与表头 ----------
    for a, b in cfg["merge_ranges"]:
        sh.merge_cells(f"{a}:{b}")
    for addr, val in cfg["set_value_cells"].items():
        sh[addr] = val

    for i, title in enumerate(cfg["headers_K_row4"]):  # K4 起表头+填充
        col_letter = chr(ord("K") + i)
        cell = sh[f"{col_letter}4"]
        cell.value = title
        cell.fill = PatternFill(start_color=cfg["header_fill_color"],
                                end_color=cfg["header_fill_color"], fill_type="solid")

    # ---------- 提取编号写入C列 ----------
    for row in sh.iter_rows(min_row=2, max_row=last_empty_row - 1, max_col=3):
        raw = str(row[1].value).strip() if row[1].value else ""
        m = re.search(r"\d+", raw)
        if m:
            row[2].value = m.group()[-5:].zfill(5)
    sh[cfg["number_header_cell"]] = "编号"

    # ---------- 从“第一页”筛选生成“第一页副本” ----------
    if cfg["src_sheet_for_copy"] in wb.sheetnames:
        s1 = wb[cfg["src_sheet_for_copy"]]
        header = [c.value for c in s1[1]]
        if cfg["warehouse_col_name"] in header:
            idx = header.index(cfg["warehouse_col_name"])
            rows = [r for r in s1.iter_rows(min_row=2, values_only=True)
                    if r[idx] == cfg["warehouse_keep_value"]]
            s_copy = wb.create_sheet(cfg["copy_sheet_name"])
            s_copy.append(header)
            for r in rows:
                s_copy.append(r)

    # ---------- 生成“家里库存” ----------
    if cfg["copy_sheet_name"] in wb.sheetnames:
        s_copy = wb[cfg["copy_sheet_name"]]
        header = [c.value for c in s_copy[1]]
        if cfg["home_name_col"] in header:
            name_idx = header.index(cfg["home_name_col"])
            qty_idx = header.index(cfg["home_qty_col"]) if cfg["home_qty_col"] in header else None

            s_home = wb.create_sheet(cfg["home_sheet_name"])
            s_home.append(cfg["home_cols"])

            for row in s_copy.iter_rows(min_row=2, values_only=True):
                name = str(row[name_idx]).strip() if row[name_idx] else ""
                code5 = name[:5] if not all('\u4e00' <= ch <= '\u9fa5' for ch in name[:5]) else ""
                qty = row[qty_idx] if qty_idx is not None else None
                if isinstance(qty, str):
                    qty = float(qty) if qty.replace(".", "", 1).isdigit() else None
                s_home.append([code5, name, qty])

            for r in s_home.iter_rows(min_row=2, max_row=s_home.max_row, min_col=3, max_col=3):
                for c in r:
                    if c.value is not None:
                        c.number_format = "#,##0.00"

    # ---------- 回填到库存表 M列（逐条打印匹配） ----------
    if cfg["home_sheet_name"] in wb.sheetnames:
        s_home = wb[cfg["home_sheet_name"]]
        tgt_col = cfg["backfill_target_col_index"]  # 13 = M

        # 构建映射：key -> (该行的 Cell 列表, 行号)
        map4, map5 = {}, {}
        for r in sh.iter_rows(min_row=2, max_row=last_empty_row - 1, max_col=tgt_col):
            cval = r[2].value  # C列（编号）
            if not cval:
                continue
            s = str(cval)
            if len(s) >= 4:
                map4[s[-4:]] = (r, r[0].row)      # 后4位映射
            map5[s.zfill(5)] = (r, r[0].row)       # 5位映射

        cnt_4 = cnt_5 = cnt_miss = 0

        for idx, row in enumerate(s_home.iter_rows(min_row=2, values_only=True), start=2):
            raw = str(row[0]).strip() if row[0] else ""
            name = row[1]
            qty = row[2]

            if re.fullmatch(cfg["regex_4digit_dash"], raw):
                k = raw[:4]
                if k in map4:
                    cells, rownum = map4[k]
                    c_val = cells[2].value
                    sh.cell(row=rownum, column=tgt_col).value = qty
                    print(f"✅ 回填(后4位匹配) 源行{idx} [{raw} | {name}] 数量={qty} → 目标行{rownum} (C={c_val}) → M{rownum}")
                    cnt_4 += 1
                else:
                    print(f"❔ 未匹配(后4位) 源行{idx} [{raw} | {name}]")
                    cnt_miss += 1

            elif re.fullmatch(cfg["regex_5digit"], raw):
                k = raw.zfill(5)
                if k in map5:
                    cells, rownum = map5[k]
                    c_val = cells[2].value
                    sh.cell(row=rownum, column=tgt_col).value = qty
                    print(f"✅ 回填(5位匹配)  源行{idx} [{raw} | {name}] 数量={qty} → 目标行{rownum} (C={c_val}) → M{rownum}")
                    cnt_5 += 1
                else:
                    print(f"❔ 未匹配(5位)   源行{idx} [{raw} | {name}]")
                    cnt_miss += 1
            else:
                print(f"⏭️ 跳过(格式不符) 源行{idx} [{raw} | {name}]")
                cnt_miss += 1

        print(f"📊 回填汇总：后4位匹配 {cnt_4} 条，5位匹配 {cnt_5} 条，未命中/跳过 {cnt_miss} 条。")

    # ---------- 会计格式与右对齐（G~Q） ----------
    c1, c2 = cfg["acc_fmt_cols"]
    for col in range(c1, c2 + 1):
        letter = get_column_letter(col)
        for r in range(cfg["acc_fmt_row_from"], last_empty_row + 1):
            cell = sh[f"{letter}{r}"]
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = cfg["acc_number_format"]

    # ---------- 指定列左对齐（如 T 列） ----------
    for col_letter in cfg["left_align_cols"]:
        for cell in sh[col_letter]:
            cell.alignment = Alignment(horizontal="left")

    # ---------- 列宽与隐藏 ----------
    for col, w in cfg["column_widths"].items():
        sh.column_dimensions[col].width = w + 0.6
    for col in cfg["hidden_columns"]:
        sh.column_dimensions[col].hidden = True

    # ---------- 视图：冻结/缩放/隐藏行 ----------
    sh.row_dimensions[1].height = cfg["row1_height"]
    sh.freeze_panes = cfg["freeze_panes"]
    sh.row_dimensions[2].outlineLevel = 1
    sh.row_dimensions[2].hidden = bool(cfg["hide_row2"])
    sh.sheet_properties.outlinePr.summaryBelow = True
    sh.sheet_view.zoomScale = cfg["zoom_scale"]

    # ---------- 聚焦：激活目标表 + 选中 focus_cell ----------
    wb.active = wb.sheetnames.index(sh.title)
    sh.sheet_view.selection = [Selection(activeCell=cfg["focus_cell"], sqref=cfg["focus_cell"])]

    # ---------- 最后一步：写入“等待您查看”的收到时间到 M3 ----------
    waiting_time = _read_waiting_time(folder_path, cfg["meta_filename"], cfg["meta_waiting_key"])
    if waiting_time:
        sh["M3"].value = waiting_time
        sh["M3"].alignment = Alignment(horizontal="left", vertical="center")
        print(f"🕒 已写入 M3（等待您查看时间）: {waiting_time}")
    else:
        print("🕒 没有可写入的等待时间（mail_meta.json 缺失或键为空）")

    # ---------- 保存 ----------
    wb.save(latest_file)
    wb.close()
    print(f"🎉 已完成处理并保存: {latest_file}")


if __name__ == "__main__":
    main(CONFIG)
