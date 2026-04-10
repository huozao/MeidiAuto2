# ================================================
# STEP CARD
# 功能: 按需求表回填 K/N/P/T 列并统一格式。
# 输入: list.xlsx, 总库存*.xlsx
# 输出: 更新后的总库存*.xlsx
# 上游: 032 Warehousing at out.py
# 下游: 041/042/050
# ================================================

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from pipeline.io_utils import find_first_excel, find_required_file, resolve_data_dir


DEMAND_XLSX = "list.xlsx"
DEMAND_SHEET = "2503"
INV_SHEET = "库存表"

START_ROW = 5
WRITE_MAP = {11: "B", 14: "C", 16: "D", 20: "E"}  # K,N,P,T <- B,C,D,E
ALIGN_COL_RANGE = (11, 20)
BORDER_ROWS = (4, 53)
BORDER_COLS = (11, 20)
FONT7_COLS = [11, 14]
FONT7_ROWS = (4, 54)


def configure_stdio() -> None:
    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        if stream is None or not hasattr(stream, "reconfigure"):
            continue
        try:
            stream.reconfigure(errors="replace")
        except Exception:
            pass


def resolve_support_dirs(data_dir: Path) -> list[Path]:
    return [data_dir, Path(__file__).resolve().parent / "data"]


def has_cn(value: object) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", str(value)))


def align_range(sheet, start_row: int, c1: int, c2: int, horizontal: str = "right") -> None:
    for col in range(c1, c2 + 1):
        for row in sheet.iter_rows(min_row=start_row, min_col=col, max_col=col):
            for cell in row:
                cell.alignment = Alignment(horizontal=horizontal)


def load_demand_data(demand_file: str) -> dict[str, tuple[object, object, object, object]]:
    wb_demand = openpyxl.load_workbook(demand_file, data_only=True)
    if DEMAND_SHEET not in wb_demand.sheetnames:
        raise ValueError(f"需求文件缺少工作表: {DEMAND_SHEET}")

    sheet_demand = wb_demand[DEMAND_SHEET]
    demand_data: dict[str, tuple[object, object, object, object]] = {}
    for a, b, c, d, e in sheet_demand.iter_rows(min_row=2, max_col=5, values_only=True):
        if a is None:
            continue
        key = str(a).strip()
        if key:
            demand_data[key] = (b, c, d, e)
    wb_demand.close()
    return demand_data


def apply_inventory_updates(inventory_file: str, demand_data: dict[str, tuple[object, object, object, object]]) -> int:
    wb_inventory = openpyxl.load_workbook(inventory_file)
    if INV_SHEET not in wb_inventory.sheetnames:
        wb_inventory.close()
        raise ValueError(f"库存文件缺少工作表: {INV_SHEET}")

    sheet_inventory = wb_inventory[INV_SHEET]
    updated = 0
    for row in sheet_inventory.iter_rows(min_row=START_ROW, max_col=20):
        code = row[2].value
        if code is None:
            continue
        values = demand_data.get(str(code).strip())
        if not values:
            continue
        row_num = row[0].row
        for target_col, source_col in WRITE_MAP.items():
            source_index = ord(source_col) - ord("B")
            sheet_inventory.cell(row=row_num, column=target_col, value=values[source_index])
        updated += 1

    align_range(sheet_inventory, START_ROW, *ALIGN_COL_RANGE)
    for row in sheet_inventory.iter_rows(min_row=START_ROW, min_col=20, max_col=20):
        for cell in row:
            cell.alignment = Alignment(horizontal="left")

    thin = Border(
        top=Side(style="thin"),
        left=Side(style="thin"),
        right=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    r1, r2 = BORDER_ROWS
    c1, c2 = BORDER_COLS
    for rows in sheet_inventory.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for cell in rows:
            cell.border = thin
            cell.font = Font(size=10)

    fr1, fr2 = FONT7_ROWS
    for col in FONT7_COLS:
        for rows in sheet_inventory.iter_rows(min_row=fr1, max_row=fr2, min_col=col, max_col=col):
            for cell in rows:
                cell.font = Font(size=7)

    for col in [11, 14]:
        for rows in sheet_inventory.iter_rows(min_row=START_ROW, min_col=col, max_col=col):
            for cell in rows:
                if cell.value is not None and has_cn(cell.value):
                    cell.font = Font(size=5)

    wb_inventory.save(inventory_file)
    wb_inventory.close()
    return updated


def main(argv: list[str] | None = None) -> int:
    configure_stdio()
    argv = argv or sys.argv
    data_dir = resolve_data_dir(argv[1] if len(argv) >= 2 else None, "data")
    if not data_dir.exists():
        print(f"❌ 库存目录不存在: {data_dir}")
        return 1

    support_dirs = resolve_support_dirs(data_dir)
    try:
        demand_file = find_required_file(support_dirs, DEMAND_XLSX, label="需求文件")
    except FileNotFoundError as exc:
        print(f"❌ {exc}")
        return 1

    inventory_file = find_first_excel(data_dir, "*总库存*.xlsx")
    if not inventory_file:
        print("❌ 未找到包含“总库存”的文件")
        return 1

    try:
        demand_data = load_demand_data(demand_file)
        updated = apply_inventory_updates(inventory_file, demand_data)
    except Exception as exc:
        print(f"❌ 处理失败: {exc}")
        return 1

    print(f"✅ 更新 {updated} 行 | 需求文件: {Path(demand_file).name} | 库存文件: {Path(inventory_file).name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
