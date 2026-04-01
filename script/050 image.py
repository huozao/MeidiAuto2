from __future__ import annotations

import glob
import os
import re
import sys
from datetime import datetime

import matplotlib.pyplot as plt
from matplotlib import font_manager, rcParams
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


# 尝试使用常见中文字体，避免邮件图中文字发虚/方块
CANDIDATE_FONTS = ["Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "PingFang SC"]
for name in CANDIDATE_FONTS:
    try:
        font_manager.findfont(name, fallback_to_default=False)
        rcParams["font.family"] = name
        break
    except Exception:
        continue


def resolve_inventory_folder(argv: list[str] | None = None) -> str:
    default_inventory_folder = os.path.abspath(os.path.join(os.getcwd(), "data"))
    argv = argv or sys.argv

    # 兼容 argv 中混入参数标记（例如 --data-dir）
    positional = [item for item in argv[1:] if not item.startswith("-")]
    if positional:
        inventory_folder = positional[-1]
        print(f"✅ 使用传入路径: {inventory_folder}")
    else:
        inventory_folder = default_inventory_folder
        print(f"⚠️ 未传入路径，使用默认路径: {inventory_folder}")

    if not os.path.exists(inventory_folder):
        print(f"❌ 文件夹路径不存在: {inventory_folder}")
        raise SystemExit(1)
    return inventory_folder


def pick_inventory_file(inventory_folder: str) -> str:
    files = glob.glob(os.path.join(inventory_folder, "总库存*.xlsx"))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        print("❌ 没有找到符合条件的总库存文件！")
        raise SystemExit(1)
    latest_file = max(files, key=os.path.getctime)
    print(f"✅ 找到最新文件：{latest_file}")
    return latest_file


def _rgb(color) -> str | None:
    if not color:
        return None
    if color.type == "rgb" and color.rgb:
        return color.rgb[-6:]
    return None


def _to_display(v) -> str:
    if v is None:
        return ""
    return str(v)


def _is_non_empty(v) -> bool:
    if v is None:
        return False
    if isinstance(v, str) and not v.strip():
        return False
    return True


def _parse_col_range(value: str) -> tuple[int, int]:
    """解析 A:T / A- T / A~T 形式。"""
    text = value.strip().upper().replace(" ", "")
    m = re.fullmatch(r"([A-Z]+)[:\-~]([A-Z]+)", text)
    if not m:
        raise ValueError(f"列范围格式错误: {value}（示例: A:T）")
    c1 = column_index_from_string(m.group(1))
    c2 = column_index_from_string(m.group(2))
    if c1 > c2:
        c1, c2 = c2, c1
    return c1, c2


def detect_used_bounds(ws, col_start: int, col_end: int, scan_from_row: int = 1) -> tuple[int, int]:
    """在指定列范围内，自动检测首尾有效行。"""
    first: int | None = None
    last: int | None = None

    for r in range(scan_from_row, ws.max_row + 1):
        has_value = any(_is_non_empty(ws.cell(row=r, column=c).value) for c in range(col_start, col_end + 1))
        if has_value and first is None:
            first = r
        if has_value:
            last = r

    if first is None or last is None:
        # 回退一个可渲染最小区间
        return 1, min(ws.max_row, 40)

    # 留一点上下边距，更接近 Excel 截图体验
    first = max(1, first - 1)
    last = min(ws.max_row, last + 2)
    return first, last


def build_table_payload(ws, col_start: int, col_end: int, row_start: int, row_end: int):
    cols = list(range(col_start, col_end + 1))
    rows = list(range(row_start, row_end + 1))

    data: list[list[str]] = []
    txt_colors: list[list[str | None]] = []
    bg_colors: list[list[str | None]] = []

    for r in rows:
        row_data: list[str] = []
        row_txt: list[str | None] = []
        row_bg: list[str | None] = []
        for c in cols:
            cell = ws.cell(row=r, column=c)
            row_data.append(_to_display(cell.value))
            row_txt.append(_rgb(cell.font.color))
            row_bg.append(_rgb(cell.fill.fgColor))
        data.append(row_data)
        txt_colors.append(row_txt)
        bg_colors.append(row_bg)

    col_widths: list[float] = []
    for c in cols:
        letter = get_column_letter(c)
        width = ws.column_dimensions[letter].width or 8.0
        col_widths.append(float(width))

    return data, txt_colors, bg_colors, col_widths


def render_table_image(data, txt_colors, bg_colors, col_widths, output_path: str) -> None:
    nrows = len(data)
    ncols = len(data[0]) if nrows else 0
    if nrows == 0 or ncols == 0:
        print("❌ 没有可渲染的数据")
        raise SystemExit(1)

    # 第二阶段优化：按有效区自动裁剪后再渲染
    fig_w = max(20, ncols * 1.35)
    fig_h = max(10, nrows * 0.50)

    fig, ax = plt.subplots(figsize=(fig_w, fig_h), dpi=220)
    ax.axis("off")

    table = ax.table(cellText=data, cellLoc="center", loc="center")
    table.auto_set_font_size(False)
    table.set_fontsize(10)

    total = sum(col_widths) if sum(col_widths) > 0 else ncols

    for r in range(nrows):
        for c in range(ncols):
            cell = table[r, c]
            cell.set_width((col_widths[c] / total) * 0.98)
            cell.set_height(0.98 / nrows)
            cell.set_linewidth(0.8)

            fg = txt_colors[r][c]
            if fg and fg != "000000":
                try:
                    cell.get_text().set_color(f"#{fg}")
                except Exception:
                    pass

            bg = bg_colors[r][c]
            if bg and bg not in ("000000", "FFFFFF"):
                try:
                    cell.set_facecolor(f"#{bg}")
                except Exception:
                    pass

    fig.savefig(output_path, bbox_inches="tight", pad_inches=0.04)
    plt.close(fig)


def main(argv: list[str] | None = None) -> int:
    folder = resolve_inventory_folder(argv or sys.argv)
    latest_file = pick_inventory_file(folder)

    wb = load_workbook(latest_file, data_only=False)
    ws = wb["库存表"] if "库存表" in wb.sheetnames else wb.active

    # 可通过环境变量覆盖导图列范围（默认 A:T）
    col_range = os.getenv("MAIL_IMAGE_COL_RANGE", "A:T")
    col_start, col_end = _parse_col_range(col_range)
    row_start, row_end = detect_used_bounds(ws, col_start=col_start, col_end=col_end)

    print(
        f"🖼️ 导图区域: {get_column_letter(col_start)}{row_start}:{get_column_letter(col_end)}{row_end}"
    )

    data, txt_colors, bg_colors, col_widths = build_table_payload(
        ws,
        col_start=col_start,
        col_end=col_end,
        row_start=row_start,
        row_end=row_end,
    )

    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    image_filepath = os.path.join(folder, f"美的仓储自动化_{current_time}.png")
    render_table_image(data, txt_colors, bg_colors, col_widths, image_filepath)

    print(f"✅ 图片已保存：{image_filepath}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
