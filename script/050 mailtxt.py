# ================================================
# STEP CARD
# 功能: 生成邮件正文 HTML（异常信息与摘要）。
# 输入: 总库存*.xlsx
# 输出: output.html
# 上游: 042 Color display.py
# 下游: 051 Send an email.py
# ================================================

import os
import sys
import glob
import openpyxl
from datetime import datetime
import re


# ================================
# 📂 配置文件路径
# ================================
def get_inventory_folder():
    default_inventory_folder = os.path.abspath(os.path.join(os.getcwd(), "data"))
    inventory_folder = sys.argv[1] if len(sys.argv) >= 2 else default_inventory_folder
    print(f"✅ 使用传入路径: {inventory_folder}" if len(
        sys.argv) >= 2 else f"⚠️ 未传入路径，使用默认路径: {inventory_folder}")

    if not os.path.exists(inventory_folder):
        print(f"❌ 文件夹路径不存在: {inventory_folder}")
        sys.exit(1)

    print(f"📂 当前工作文件夹: {inventory_folder}")
    return inventory_folder


# ================================
# 1. 查找 Excel 文件
# ================================
def find_excel_file(inventory_folder):
    files = glob.glob(os.path.join(inventory_folder, '总库存*.xlsx'))
    valid_files = [f for f in files if not os.path.basename(f).startswith('~$')]

    if not valid_files:
        print("❌ 没有找到符合条件的文件！")
        sys.exit(1)

    print(f"✅ 找到文件：{valid_files[0]}")
    return valid_files[0]


# ================================
# 2. 读取工作表
# ================================
def load_worksheet(inventory_file, sheet_name="库存表"):
    try:
        wb = openpyxl.load_workbook(inventory_file)
    except Exception as e:
        print(f"❌ 无法打开 Excel 文件：{e}")
        sys.exit(1)

    if sheet_name not in wb.sheetnames:
        print(f"❌ 工作表“{sheet_name}”不存在！")
        sys.exit(1)

    return wb[sheet_name]


# ================================
# 🎨 颜色判断函数（只识别填充色）
# ================================
def get_cell_fill_rgb(cell):
    fill = cell.fill
    if fill and fill.fill_type == "solid":
        color = fill.fgColor
        if color.type == "rgb" and color.rgb:
            return color.rgb[-6:].upper()
    return None


def is_fill_color(cell, color_code: str):
    return get_cell_fill_rgb(cell) == color_code.upper()


# ================================
# 查找红色或紫色的行
# ================================
def find_colored_rows(sheet):
    colored_rows = []
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=12)
        if is_fill_color(cell, "FF0000") or is_fill_color(cell, "3F0065"):
            colored_rows.append(row)
            color_hex = get_cell_fill_rgb(cell)
            print(f"✅ 符合条件颜色 → 行: {row} RGB: #{color_hex}")
    return colored_rows


# ================================
# 📅 获取日期（H3 与 M3）
# ================================
def _fmt_dt(v):
    """把单元格时间或字符串格式化为 'YYYY-MM-DD HH:MM:SS'；无法解析就原样返回/空串。"""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass
        try:
            # ISO8601 兜底，如 2025-09-19T18:00:05 或带Z
            return datetime.fromisoformat(s.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return s
    return ""


def get_dates(sheet):
    """
    返回 (date, date2)
    - date  来自 H3（原标题用）
    - date2 来自 M3（家里库存数据用）
    """
    date = _fmt_dt(sheet["H3"].value)
    date2 = _fmt_dt(sheet["M3"].value)
    return date, date2


# ================================
# 查找 B 列第一个空单元格
# ================================
def find_last_empty_row(sheet):
    for row in range(4, sheet.max_row + 1):
        if sheet[f"B{row}"].value is None:
            return row
    return sheet.max_row + 1


# ================================
# 计算公式并返回总和
# ================================
def calculate_sum(sheet, formula):
    if isinstance(formula, (int, float)):
        return formula
    if not isinstance(formula, str):
        return 0

    match = re.match(r"^=SUM\((.+)\)$", formula)
    if match:
        cell_range = match.group(1)
        start_cell, end_cell = cell_range.split(":")
        start_row, start_col = int(start_cell[1:]), openpyxl.utils.cell.column_index_from_string(start_cell[:1])
        end_row, end_col = int(end_cell[1:]), openpyxl.utils.cell.column_index_from_string(end_cell[:1])

        total = 0
        for row in range(start_row, end_row + 1):
            value = sheet.cell(row=row, column=start_col).value
            if isinstance(value, (int, float)):
                total += value
        return total
    return 0


# ================================
# 获取库存合计信息
# ================================
def prepare_summary_text(sheet, last_empty_row):
    stock_total       = calculate_sum(sheet, sheet.cell(row=last_empty_row, column=10).value)
    monthly_plan      = calculate_sum(sheet, sheet.cell(row=last_empty_row, column=16).value)
    plan_gap_output   = calculate_sum(sheet, sheet.cell(row=last_empty_row, column=17).value)
    monthly_sent      = calculate_sum(sheet, sheet.cell(row=last_empty_row, column=18).value)
    monthly_received  = calculate_sum(sheet, sheet.cell(row=last_empty_row, column=19).value)
    monthly_remaining = monthly_plan - monthly_sent if monthly_plan and monthly_sent else 0

    print(f"📊 库存总量: {stock_total}, 月计划: {monthly_plan}, 缺口排产: {plan_gap_output}, 出库: {monthly_sent}, 入库: {monthly_received}")
    return stock_total, monthly_plan, plan_gap_output, monthly_sent, monthly_received, monthly_remaining


# ================================
# 构建输出文本
# ================================
def construct_html_content(sheet, colored_rows, date, date2,
                           stock_total, monthly_plan, plan_gap_output,
                           monthly_sent, monthly_received, monthly_remaining):
    html = """
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            table { border-collapse: collapse; width: auto; margin-top: 10px; }
            th, td { border: 1px solid #999; padding: 6px 10px; }
            th { background-color: #f2f2f2; text-align: left; }
            td.right { text-align: right; }
            td.left { text-align: left; }
        </style>
    </head>
    <body>
    """

    # 两个标题：H3 对应重庆俊都仓储，M3 对应家里库存
    html += f"""
    <h1>{date} 重庆俊都仓储数据</h1>
    <h1>{date2} 家里库存数据</h1>
    <h5>“外仓库存＜50%外仓应存数量”的物料有 <strong>{len(colored_rows)}</strong> 款</h5>
    """

    html += """
    <table>
        <tr>
            <th>编号</th>
            <th>库存</th>
            <th>外应存</th>
            <th>家里库存</th>
        </tr>
    """
    for row in colored_rows:
        code = sheet.cell(row=row, column=3).value
        stock = sheet.cell(row=row, column=10).value
        expected = sheet.cell(row=row, column=11).value
        home_stock = sheet.cell(row=row, column=13).value

        stock_fmt = f"{stock:,.1f}" if isinstance(stock, (int, float)) else stock
        expected_fmt = f"{expected:,.1f}" if isinstance(expected, (int, float)) else expected
        home_stock_fmt = f"{home_stock:,.1f}" if isinstance(home_stock, (int, float)) else home_stock

        html += f"""
        <tr>
            <td>{code}</td>
            <td class="right">{stock_fmt}</td>
            <td class="right">{expected_fmt}</td>
            <td class="right">{home_stock_fmt}</td>
        </tr>
        """

    html += "</table>"

    html += """
    <h5>汇总信息</h5>
    <table>
        <tr><th>项目</th><th>数值</th></tr>
    """

    def row(label, value):
        return f"""
        <tr>
            <td class="left">{label}</td>
            <td class="right">{value:,.1f}</td>
        </tr>
        """

    html += row("外仓库存总量", stock_total)
    html += row("月计划", monthly_plan)
    html += row("月计划缺口排产", plan_gap_output)
    html += row("外仓出库总量", monthly_sent)
    html += row("外仓入库总量", monthly_received)
    html += row("月预估还有要发货", monthly_remaining)

    html += "</table>\n</body></html>"
    return html


# ================================
# 保存为 HTML 文件
# ================================
def save_output_to_file(html_content, output_dir):
    output_filename = os.path.join(output_dir, "output.html")
    with open(output_filename, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"📁 已成功保存为 HTML 文件：{output_filename}")


# ================================
# 主函数
# ================================
def main(argv: list[str] | None = None) -> int:
    if argv is not None:
        sys.argv = argv
    inventory_folder = get_inventory_folder()
    inventory_file = find_excel_file(inventory_folder)
    sheet = load_worksheet(inventory_file)

    colored_rows = find_colored_rows(sheet)
    date, date2 = get_dates(sheet)  # 👈 同时拿 H3 / M3
    last_empty_row = find_last_empty_row(sheet)
    print(f"⚡ 发现 B 列第一个空单元格所在行: {last_empty_row}")

    stock_total, monthly_plan, plan_gap_output, monthly_sent, monthly_received, monthly_remaining = prepare_summary_text(sheet, last_empty_row)

    html_content = construct_html_content(
        sheet, colored_rows, date, date2,
        stock_total, monthly_plan, plan_gap_output,
        monthly_sent, monthly_received, monthly_remaining
    )

    print("\n📋 HTML 已生成，预览内容省略…")
    save_output_to_file(html_content, inventory_folder)
    print("✅ 红色或紫色单元格数量：", len(colored_rows))
    print("📌 行号列表：", colored_rows)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
